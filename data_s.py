import openpyxl
from datetime import datetime

book1 = openpyxl.load_workbook("cours1.xlsx")
book2 = openpyxl.load_workbook("cours2.xlsx")
book3 = openpyxl.load_workbook("cours3.xlsx")
sheet1 = book1.active
sheet2 = book2.active
sheet3 = book3.active

d___d = ['вчера', 'сегодня', 'завтра']

days = ["ПОНЕДЕЛЬНИК", "ВТОРНИК", "СРЕДА", "ЧЕТВЕРГ", "ПЯТНИЦА", "СУББОТА", "ВОСКРЕСЕНЬЕ"]

today = datetime.today().weekday()

max_cols1 = sheet1.max_column
max_row1 = sheet1.max_row

max_cols2 = sheet2.max_column
max_row2 = sheet2.max_row

max_cols3 = sheet3.max_column
max_row3 = sheet3.max_row
def rasp(key, text):
    if key == "ВОСКРЕСЕНЬЕ":
        return "NO"

    result = []
    flag = False

    if text[-2:] == "23":
        group = sheet1.cell(row=2, column=6)
        rw = 2
        col = 6

        # Finding the correct column for the group
        while col <= max_cols1:
            if sheet1.cell(row=rw, column=col).value == text:
                flag = True
                break
            else:
                col += 5

        if not flag:
            return "NO"

        flag = False
        rw = 4  # Reset rw to start checking rows

        # Finding the correct row for the day
        while rw <= max_row1:
            if sheet1.cell(row=rw, column=1).value == key:
                flag = True
                break
            else:
                rw += 14

        if not flag:
            return "NO"

        # Collecting schedule details
        for i in range(14):
            row = rw + i
            if i % 2 == 0:
                day = "I"
            else:
                day = "II"
            subject = sheet1.cell(row=row, column=col).value
            if "Физика" in subject:
                subject = subject[:len(subject)//2 -1]
            result.append(f"- {day} - {subject}")

        return result

    elif text[-2:] == "22":
        group = sheet2.cell(row=2, column=6)
        rw = 2
        col = 6

        # Finding the correct column for the group
        while col <= max_cols2:
            if sheet2.cell(row=rw, column=col).value == text:
                flag = True
                break
            else:
                col += 5

        if not flag:
            return "NO"

        flag = False
        rw = 4  # Reset rw to start checking rows

        # Finding the correct row for the day
        while rw <= max_row2:
            if sheet2.cell(row=rw, column=1).value == key:
                flag = True
                break
            else:
                rw += 14

        if not flag:
            return "NO"

        # Collecting schedule details
        for i in range(14):
            row = rw + i
            if i % 2 == 0:
                day = "I"
            else:
                day = "II"
            subject = sheet2.cell(row=row, column=col).value
            if "Физика" in subject:
                subject = subject[:len(subject) // 2 - 1]
            result.append(f"- {day} - {subject}")

        return result

    elif text[-2:] == "21":
        group = sheet3.cell(row=2, column=6)
        rw = 2
        col = 6

        # Finding the correct column for the group
        while col <= max_cols3:
            if sheet3.cell(row=rw, column=col).value == text:
                flag = True
                break
            else:
                col += 5

        if not flag:
            return "NO"

        flag = False
        rw = 4  # Reset rw to start checking rows

        # Finding the correct row for the day
        while rw <= max_row3:
            if sheet3.cell(row=rw, column=1).value == key:
                flag = True
                break
            else:
                rw += 14

        if not flag:
            return "NO"

        # Collecting schedule details
        for i in range(14):
            row = rw + i
            if i % 2 == 0:
                day = "I"
            else:
                day = "II"
            subject = sheet3.cell(row=row, column=col).value
            if "Физика" in subject:
                subject = subject[:len(subject) // 2 - 1]
            result.append(f"- {day} - {subject}")

        return result
    else:
        return "NO"


# for i in rasp("ПОНЕДЕЛЬНИК", "ИКБО-03-22"):
#     print(i)

prepod = []
def prepods():
    for j in range(8,448,10):
        for i in range(4,87):
            cell = sheet1.cell(row = i , column = j).value
            if type(cell) == str:
                if cell.count ("п/г")>0:
                    cell = cell.split("п/г")
                    prepod.append(cell[0])
                    prepod.append(cell[1])
                else:
                    prepod.append(cell)
    return prepod

prepod = prepods()
print(set(prepod))

def rasp_prepods(key , name , chet = 0):
    for i in range(len(d___d)):
        if key.lower() ==  d___d[i]:
            if today == 6 and i == 1:
                return 'no'
            if today == 5 and i == 2:
                return 'no'
            if today == 0 and i == 0:
                return 'no'

    if key.lower() == 'сегодня':
        key = days[datetime.today().weekday()]
    elif key.lower() == 'вчера':
        key = days[datetime.today().weekday()-1]
    elif key.lower() == 'завтра':
        if today == 6:
            key = days[0]
        else:
            key = days[datetime.today().weekday()+1]

    b = 1
    for j in range(8, 448, 5):
        for i in range(4 + chet, 87, 2):
            if (i - 4) % 14 == 0:
                b = 0
            b += 1

            cell = sheet1.cell(row=i, column=j).value
            if type(cell) == str:
                cell = cell.lower()
                if cell.count(name) > 0:
                    if days[int((i - 4) // 14)] == key:
                        yield (b, int((i - 4) // 14), sheet1.cell(row=i, column=j - 2).value,
                               sheet1.cell(row=i, column=j - 2).value)

for i in rasp_prepods('СРЕДА','вервальд', 0):
    print(i)