import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard , VkKeyboardColor
from vk_api import VkUpload
from datetime import datetime
from pyowm.utils import config as cfg
import requests
from bs4 import BeautifulSoup
from data_s import prepods , rasp_prepods
import openpyxl
import PIL.Image as Image
import json
import pyowm



def degrees_to_direction(degrees):
    directions = ["Север", "Северо-Северо-Восток", "Северо-Восток", "Восток-Северо-Восток",
                  "Восток", "Восток-Юго-Восток", "Юго-Восток", "Юго-Юго-Восток",
                  "Юг", "Юго-Юго-Запад", "Юго-Запад", "Запад-Юго-Запад",
                  "Запад", "Запад-Северо-Запад", "Северо-Запад", "Северо-Северо-Запад"]
    index = int((degrees / 22.5) + 0.5) % 16
    return directions[index]

def main():

    def rasp(key, text):
        if key == "ВОСКРЕСЕНЬЕ":
            return "NO"

        result = []
        flag = False

        if text[-2:] == "23":
            group = sheet1.cell(row=2, column=6)
            rw = 2
            col = 6

            # Finding the correct column for the group
            while col <= max_cols1:
                if sheet1.cell(row=rw, column=col).value == text:
                    flag = True
                    break
                else:
                    col += 5

            if not flag:
                return "NO"

            flag = False
            rw = 4  # Reset rw to start checking rows

            # Finding the correct row for the day
            while rw <= max_row1:
                if sheet1.cell(row=rw, column=1).value == key:
                    flag = True
                    break
                else:
                    rw += 14

            if not flag:
                return "NO"

            # Collecting schedule details
            for i in range(14):
                row = rw + i
                if i % 2 == 0:
                    day = "I"
                else:
                    day = "II"
                subject = sheet1.cell(row=row, column=col).value
                if "Физика" in subject:
                    subject = subject[:len(subject)]
                result.append(f"- {day} - {subject}")

            return result

        elif text[-2:] == "22":
            group = sheet2.cell(row=2, column=6)
            rw = 2
            col = 6

            # Finding the correct column for the group
            while col <= max_cols2:
                if sheet2.cell(row=rw, column=col).value == text:
                    flag = True
                    break
                else:
                    col += 5

            if not flag:
                return "NO"

            flag = False
            rw = 4  # Reset rw to start checking rows

            # Finding the correct row for the day
            while rw <= max_row2:
                if sheet2.cell(row=rw, column=1).value == key:
                    flag = True
                    break
                else:
                    rw += 14

            if not flag:
                return "NO"

            # Collecting schedule details
            for i in range(14):
                row = rw + i
                if i % 2 == 0:
                    day = "I"
                else:
                    day = "II"
                subject = sheet2.cell(row=row, column=col).value
                if "Физика" in subject:
                    subject = subject[:len(subject)]
                result.append(f"- {day} - {subject}")

            return result

        elif text[-2:] == "21":
            group = sheet3.cell(row=2, column=6)
            rw = 2
            col = 6

            # Finding the correct column for the group
            while col <= max_cols3:
                if sheet3.cell(row=rw, column=col).value == text:
                    flag = True
                    break
                else:
                    col += 5

            if not flag:
                return "NO"

            flag = False
            rw = 4  # Reset rw to start checking rows

            # Finding the correct row for the day
            while rw <= max_row3:
                if sheet3.cell(row=rw, column=1).value == key:
                    flag = True
                    break
                else:
                    rw += 14

            if not flag:
                return "NO"

            # Collecting schedule details
            for i in range(14):
                row = rw + i
                if i % 2 == 0:
                    day = "I"
                else:
                    day = "II"
                subject = sheet3.cell(row=row, column=col).value
                if "Физика" in subject:
                    subject = subject[:len(subject) ]
                result.append(f"- {day} - {subject}")

            return result
        else:
            return "NO"

    def change_day():
        day_map = {
        0 : "pn.png",
        1 : "vt.png",
        2 : "sr.png",
        3 : "cht.png",
        4 : "pt.png",
        5 : "sb.png"
        }
        current_day = datetime.today().weekday()
        return day_map.get(current_day,"Выходной")

    def weather_today():
        s_city = "Moscow,RU"
        city_id = 0
        appid = "24665a75e220a590bc39deaccbfc1a78"
        try:
            res = requests.get("http://api.openweathermap.org/data/2.5/find",
                               params={'q': s_city, 'type': 'like', 'units': 'metric', 'APPID': appid})
            data = res.json()

            print("Ответ данных :", data)
            if "list" in data and len(data["list"]) > 0:
                cities = ["{} ({})".format(d['name'], d['sys']['country'])
                          for d in data['list']]
                print("Города : ", cities)
                city_id = data["list"][0]["id"]
                print("city_id = ", city_id)
            else:
                print("Города не найдены")
        except Exception as e:
            print("Exception (find):", e)
            pass

        try:
            res = requests.get("http://api.openweathermap.org/data/2.5/weather",
                               params={'id': city_id, 'units': 'metric', 'lang': 'ru', 'APPID': appid})
            data = res.json()
            print("Погода:", data['weather'][0]['description'])
            print("Текущая температура:", data['main']['temp'])
            print("Минимальная температура:", data['main']['temp_min'])
            print("Максимальная температура:", data['main']['temp_max'])
            print("Давление:", data["main"]["pressure"], "мм рт.ст.")
            print("Влажность:", data["main"]["humidity"], "%")
            print("Скорость ветра:", data["wind"]["speed"], "м.с")
            wind_deg = data["wind"].get("deg")
            wind_deg = degrees_to_direction(wind_deg)
            print("Направление ветра:", wind_deg)

            vk.messages.send(
                user_id=event.user_id,
                random_id=get_random_id(),
                message="Информация о погоде в Москве \n"
                        "Погода: " + data['weather'][0]['description'] + "\n"
                                                                         "Текущая температура: " + str(
                    data['main']['temp']) + "°C\n"
                                            "Минимальная температура: " + str(data['main']['temp_min']) + "°C\n"
                                                                                                          "Максимальная температура: " + str(
                    data['main']['temp_max']) + "°C\n"
                                                "Давление: " + str(data["main"]["pressure"]) + "мм рт.ст.\n"
                                                                                               "Влажность: " + str(
                    data["main"]["humidity"]) + "% \n"
                                                "Скорость ветра: " + str(data["wind"]["speed"]) + "м/с\n"
                                                                                                  "Направление ветра: " + wind_deg+"\n"
                                                "Текущее время: " + str(datetime.now().strftime("%H:%M"))
            )

        except Exception as e:
            print("Exception (weather):", e)
            pass


    def weather_5_days(vk,event,owm):
        config = owm.config
        config["language"] = "ru"
        city = "Moscow,RU"
        mgr = owm.weather_manager()
        observation = mgr.weather_at_place(city)
        w = observation.weather
        forecast = mgr.forecast_at_place(city, interval="3h")

        weather_data = forecast.forecast.weathers

        weathers_json = []

        count = 0
        for weather in weather_data:
            data = {
                "date": weather.reference_time('iso'),
                "temperature": weather.temperature("celsius"),
                "status": weather.detailed_status,
            }
            count += 1

            if (count == 8):
                print("Период : " + data["date"])
                print("Температура : " + str(data["temperature"]["temp"]))
                print("Погода : " + data["status"])
                count = 0
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    message="Период : " + data["date"][:19] + "\n"
                                                         "Температура : " + str(data["temperature"]["temp"]) + "°C\n"
                                                                                                               "Погода : " +
                            data["status"] + ""
                )
            weathers_json.append(data)

        with open("weather_forecast.json", "w") as file:
            json.dump(weathers_json, file, indent=4)

        print("Прогноз погоды успешно записан")

    def br_return():
        vk.messages.send(
            user_id = event.user_id,
            random_id = get_random_id(),
            message = "Что хотите узнать ?",
            keyboard = keyboard.get_keyboard()
        )

    book1 = openpyxl.load_workbook("cours1.xlsx")
    book2 = openpyxl.load_workbook("cours2.xlsx")
    book3 = openpyxl.load_workbook("cours3.xlsx")
    print("Ready !")
    sheet1 = book1.active
    sheet2 = book2.active
    sheet3 = book3.active

    prepod = set(prepods())
    days = ["ПОНЕДЕЛЬНИК", "ВТОРНИК", "СРЕДА", "ЧЕТВЕРГ", "ПЯТНИЦА", "СУББОТА", "ВОСКРЕСЕНЬЕ"]

    max_cols1 = sheet1.max_column
    max_row1 = sheet1.max_row

    max_cols2 = sheet2.max_column
    max_row2 = sheet2.max_row

    max_cols3 = sheet3.max_column
    max_row3 = sheet3.max_row

    vk_session = vk_api.VkApi(token="vk1.a.W1EfOPCIjwDy_ZRQxqp9d1magchqv6FBRxAS8lke89YfYogE6jMaJCzGFf1aieiRMU3hlub6dybvyP4jKWRwH-HGuWbFfhVtLHrZAhpDrUvHLN_es3FVhDexpWyPmG8L6VAvO4goszxzTq-MkGxuZGZVqATLQ4ItoWTzRjEPtQVeDnt6P6FfTFzDdsmbxmfPf197Z5CsMiuqUTx3lHxGVQ")

    vk_user = vk_api.VkApi(token = "vk1.a.LkRGMw50PPxIArHZNEUhdcK03U5Di5t7kr9UwVGSIoZVFkpROa3UV8ykHfDJrc8cz7DM2bdzNdeVbu65ykoqnxJvF6H3grFIqGcPbmNm_4kWEGS1zNSPupQRZAEzK83vX8YJ5bAMvMWsRKrg-rx6s1XrCn2jAp0s9Kr_HMN4FT1fHVA4BgvTitZ6oej-zbl0")
    vk = vk_session.get_api()
    longpoll = VkLongPoll(vk_session)
    upload = VkUpload(vk_session)


    config = cfg.get_default_config()
    config["language"] = "ru"
    owm = pyowm.OWM("24665a75e220a590bc39deaccbfc1a78",config)
    user_state = {}

    keyboard = VkKeyboard()
    keyboard.add_button("Расписание" , color = VkKeyboardColor.POSITIVE)
    keyboard.add_button("Погода в Москве", color = VkKeyboardColor.NEGATIVE )

    check = False
    check2 = False

    responsed_user = set()
    count = 0

    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.text and event.to_me:
            text = event.text
            user_id = event.user_id
            if user_id not in responsed_user and count <= 0:
                print("New from {}, text = {}".format(event.user_id, event.text))
                count += 1
                vk.messages.send(
                    user_id = event.user_id,
                    random_id = get_random_id(),
                    message = "Здраствуйте , " + \
                            vk.users.get(user_id = event.user_id)[0]["first_name"] + \
                            ", с помощью кнопок ниже вы можете узнать расписание или погоду. \n"
                            "Итак, что вас интересует ?",
                    keyboard = keyboard.get_keyboard()
                    )
                responsed_user.add(event.message_id)
                check = True
                user_state[user_id] = "initial"

            elif(check == True and check2 == False and (text!= "Расписание" and text!="Погода в Москве")):
                vk.messages.send(
                    user_id = event.user_id,
                    random_id = get_random_id(),
                    message = "Я вас не понял !"
                )
                br_return()
                user_state[user_id] = "initial"

            elif(user_state[user_id ] == "initial"):
                if (text == "Погода в Москве"):
                    check2 = True
                    weather_keyboard = VkKeyboard(one_time=True)
                    weather_keyboard.add_button("Сегодня",color=VkKeyboardColor.NEGATIVE)
                    weather_keyboard.add_button("На 5 дней",color=VkKeyboardColor.POSITIVE)
                    vk.messages.send(
                        user_id = event.user_id,
                        random_id = get_random_id(),
                        message = "Выберите период:",
                        keyboard = weather_keyboard.get_keyboard()
                    )
                    user_state[user_id] = "get_weather"

                elif(text == "Расписание"):
                    check2 = True
                    key_person = VkKeyboard(one_time= True)
                    key_person.add_button("Преподаватель",color = VkKeyboardColor.PRIMARY)
                    key_person.add_button("Студент", color = VkKeyboardColor.PRIMARY)
                    vk.messages.send(
                        user_id = event.user_id,
                        random_id = get_random_id(),
                        message = "Какое расписание хотите посмотреть ? : ",
                        keyboard = key_person.get_keyboard()
                    )
                    day_s = ""
                    name_s = ""
                    user_state[user_id] = "change_person"

            elif(user_state[user_id]== "change_person"):
                if(text == "Студент"):
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="Введите группу : "
                    )
                    user_state[user_id] = "get_rasp"
                elif(text == "Преподаватель"):
                    user_state[user_id] = "get_teacher"
                    vk.messages.send(
                        user_id = event.user_id,
                        random_id = get_random_id(),
                        message = "Введите фамилию перподавателя : "
                    )
                    user_state[user_id] = "get_prepod"
                else:
                    vk.messages.send(
                        user_id = event.user_id,
                        random_id = get_random_id(),
                        message = "Я вас не понял !"
                    )
                    br_return()

            elif(user_state[user_id]=='get_weather'):
                if(text == "Сегодня"):
                    weather_today()
                    user_state[user_id] = "initial"
                    br_return()
                    check2 = False
                elif(text == "На 5 дней"):
                    weather_5_days(vk,event,owm)
                    user_state[user_id] = "initial"
                    br_return()
                    check2 = False
                else:
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="Я вас не понял !"
                    )
                    check2 = False
                    br_return()
                    user_state[user_id] = "initial"


            elif(user_state[user_id] == "get_rasp"):
                groups = text
                key_day = VkKeyboard(one_time=True)
                for i in days[:3]:
                    key_day.add_button(i , VkKeyboardColor.NEGATIVE)
                key_day.add_line()
                for i in days[3:6]:
                    key_day.add_button(i , VkKeyboardColor.POSITIVE)
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    message="Введите день : ",
                    keyboard = key_day.get_keyboard()
                )
                user_state[user_id] = "print_rasp"

            elif(user_state[user_id] == "get_prepod"):
                name_s = text
                key_day = VkKeyboard(one_time=True)
                for i in days[:3]:
                    key_day.add_button(i , VkKeyboardColor.NEGATIVE)
                key_day.add_line()
                for i in days[3:6]:
                    key_day.add_button(i , VkKeyboardColor.POSITIVE)
                vk.messages.send(
                    user_id = event.user_id,
                    random_id = get_random_id(),
                    message = "Введите день : ",
                    keyboard=key_day.get_keyboard()
                )
                user_state[user_id] = "print_rasp_prepod"

            elif(user_state[user_id] == "print_rasp_prepod"):
                day_s = text
                data_prepod_rasp = set(rasp_prepods(day_s,name_s.lower(),0))
                string = ""
                count_pars = 1
                for i in data_prepod_rasp:
                    string += str(count_pars) + " - " + str(i)[5:str(i).rfind(",")] + "\n"
                    count_pars += 1
                vk.messages.send(
                    user_id = event.user_id,
                    random_id = get_random_id(),
                    message = string
                )
                br_return()

            elif(user_state[user_id] == "print_rasp"):
                day_s = text
                data_rasp = rasp(day_s , groups)
                print(day_s , groups)
                string = ""
                for i in data_rasp:
                    string += str(i)+"\n"
                    print(i)
                vk.messages.send(
                    user_id = event.user_id,
                    random_id = get_random_id(),
                    message = string
                )
                br_return()

if __name__ == '__main__':
    main()



